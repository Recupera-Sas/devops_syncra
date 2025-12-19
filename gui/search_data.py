import pandas.errors as pd_errors
import pandas as pd
import os
from openpyxl import load_workbook
import polars as pl
from pathlib import Path
import json
import shutil
from datetime import datetime
import tempfile
import pyarrow.parquet as pq
import pyarrow as pa

def search_values_in_files(directory, output_path, search_list, process_data):
    search_values = [v.strip() for v in search_list.split(',')]
    print(f"🔎 Searching for values: {search_values}")
    
    # Create special folder for results
    base_output_path = os.path.join(output_path, "---- BUSQUEDA ----")
    os.makedirs(base_output_path, exist_ok=True)
    
    datetime_now = pd.Timestamp.now().strftime('%Y%m%d - %H%M')
    base_filename = f'Search Results {datetime_now}'
    
    results = {}
    search_stats = {
        'search_values': search_values,
        'total_files_processed': 0,
        'files_with_matches': 0,
        'total_matches': 0,
        'files_by_type': {'csv': 0, 'excel': 0, 'parquet': 0},
        'matches_by_value': {},
        'files_with_errors': []
    }
    
    print("📂 Searching through CSV, Excel and Parquet files in all subdirectories...")

    # Function to find files recursively
    def find_files_recursive(directory, extensions):
        """Find all files with specified extensions recursively"""
        file_list = []
        for extension in extensions:
            pattern = f"**/*{extension}"
            files = Path(directory).glob(pattern)
            file_list.extend([str(f) for f in files if f.is_file()])
        return file_list

    # Optimized Parquet reading with multiple fallback strategies
    def read_parquet_optimized(file_path):
        """Ultra-fast Parquet file reading with OneDrive compatibility"""
        filename = os.path.basename(file_path)
        relative_path = os.path.relpath(file_path, directory)
        
        print(f"📦 Attempting to read Parquet: {relative_path}")
        
        # Strategy 1: Try Polars with different approaches
        try:
            # Try with use_pyarrow=True option
            df_pl = pl.read_parquet(file_path, use_pyarrow=True)
            print(f"✅ Parquet loaded successfully with PyArrow: {filename} - Shape: {df_pl.shape}")
            return df_pl
        except Exception as e1:
            print(f"⚠️ Polars+PyArrow failed: {e1}. Trying direct Polars...")
            
            try:
                # Try without PyArrow
                df_pl = pl.read_parquet(file_path)
                print(f"✅ Parquet loaded successfully: {filename} - Shape: {df_pl.shape}")
                return df_pl
            except Exception as e2:
                print(f"⚠️ Direct Polars failed: {e2}. Trying PyArrow directly...")
                
                try:
                    # Strategy 2: Use PyArrow directly
                    table = pq.read_table(file_path)
                    df_pl = pl.from_arrow(table)
                    print(f"✅ Parquet loaded via direct PyArrow: {filename} - Shape: {df_pl.shape}")
                    return df_pl
                except Exception as e3:
                    print(f"⚠️ PyArrow direct failed: {e3}. Trying memory copy strategy...")
                    
                    try:
                        # Strategy 3: Copy file to temp location
                        with tempfile.NamedTemporaryFile(suffix='.parquet', delete=False) as tmp_file:
                            temp_path = tmp_file.name
                            # Copy file content
                            with open(file_path, 'rb') as src:
                                shutil.copyfileobj(src, tmp_file)
                        
                        try:
                            # Read from temp location
                            df_pl = pl.read_parquet(temp_path)
                            print(f"✅ Parquet loaded from temp copy: {filename} - Shape: {df_pl.shape}")
                            return df_pl
                        finally:
                            # Clean up temp file
                            try:
                                os.unlink(temp_path)
                            except:
                                pass
                                
                    except Exception as e4:
                        print(f"⚠️ Temp copy failed: {e4}. Trying pandas as last resort...")
                        
                        try:
                            # Strategy 4: Use pandas (slowest but most compatible)
                            df_pd = pd.read_parquet(file_path, engine='pyarrow')
                            df_pl = pl.from_pandas(df_pd)
                            print(f"✅ Parquet loaded via pandas: {filename} - Shape: {df_pl.shape}")
                            return df_pl
                        except Exception as e5:
                            print(f"💥 All Parquet reading methods failed for {filename}: {e5}")
                            search_stats['files_with_errors'].append(f"{relative_path} - Parquet error: {e5}")
                            return None

    # Rest of the CSV and Excel reading functions remain the same
    def read_csv_dynamic(file_path, delimiter=';'):
        try:
            return pd.read_csv(file_path, delimiter=delimiter, encoding='utf-8', dtype=str, low_memory=False)
        except UnicodeDecodeError:
            print(f"⚠️ Encoding issue in {file_path}. Retrying with 'latin-1'...")
            try:
                return pd.read_csv(file_path, delimiter=delimiter, encoding='latin-1', dtype=str, low_memory=False)
            except Exception as e:
                print(f"❌ Still failed with 'latin-1': {e}")
                search_stats['files_with_errors'].append(f"{file_path} - Encoding error")
                return None
        except pd_errors.EmptyDataError:
            print(f"⚠️ Empty file: {file_path}")
            return None
        except pd_errors.ParserError as e:
            print(f"❌ Parsing error in {file_path}: {e}")
            try:
                print(f"⚠️ Trying to skip bad lines in {file_path}...")
                return pd.read_csv(file_path, delimiter=delimiter, encoding='utf-8', on_bad_lines='skip', dtype=str, low_memory=False)
            except Exception as e2:
                print(f"❌ Could not recover from parse error: {e2}")
                search_stats['files_with_errors'].append(f"{file_path} - Parsing error")
                return None
        except Exception as e:
            print(f"❌ Unexpected error in {file_path}: {e}")
            search_stats['files_with_errors'].append(f"{file_path} - Unexpected error: {e}")
            return None

    def read_csv_polars_optimized(file_path, delimiter=';'):
        """Optimized CSV reading with Polars for large files"""
        try:
            df_pl = pl.read_csv(
                file_path, 
                separator=delimiter, 
                infer_schema_length=10000,
                null_values=["", "NULL", "null", "NaN", "nan"],
                try_parse_dates=False,
                dtype_backend='string'
            )
            return df_pl
        except Exception as e:
            print(f"❌ Polars failed for {file_path}: {e}. Falling back to pandas...")
            df_pd = read_csv_dynamic(file_path, delimiter)
            if df_pd is not None:
                try:
                    return pl.from_pandas(df_pd)
                except Exception as pl_error:
                    print(f"❌ Could not convert pandas to polars: {pl_error}. Using manual conversion...")
                    return convert_dataframe_safely(df_pd)
            return None

    def convert_dataframe_safely(df_pd):
        """Safely convert pandas DataFrame to Polars handling any data type issues"""
        try:
            df_pd = df_pd.astype(str)
            df_pd = df_pd.fillna('')
            return pl.from_pandas(df_pd)
        except Exception as e:
            print(f"❌ Safe conversion failed: {e}. Creating new DataFrame...")
            columns = df_pd.columns.tolist()
            data = {}
            for col in columns:
                data[col] = df_pd[col].astype(str).fillna('').tolist()
            return pl.DataFrame(data)

    def search_in_dataframe_polars(df_pl, value, filepath, sheet_name="N/A"):
        """Optimized search with Polars - CONTAINS search (not exact match)"""
        try:
            # Convert all columns to string for consistent searching and type consistency
            df_str = df_pl.cast(pl.Utf8, strict=False).fill_null("")
            
            mask = pl.lit(False)
            for col in df_str.columns:
                # Use contains for partial matching (not exact)
                col_mask = df_str[col].str.contains(value, literal=True)
                mask = mask | col_mask
            
            matching_rows = df_str.filter(mask)
            
            if len(matching_rows) > 0:
                # Preserve original columns structure
                original_columns = df_str.columns
                
                # Add metadata columns at the beginning
                matching_rows = matching_rows.with_columns([
                    pl.lit(filepath).cast(pl.Utf8).alias("Origen_Archivo"),
                    pl.lit(value).cast(pl.Utf8).alias("Dato_Buscado"),
                    pl.lit(sheet_name).cast(pl.Utf8).alias("Hoja_Origen")
                ])
                
                # Reorder columns: metadata first, then original columns
                columns_order = ["Origen_Archivo", "Dato_Buscado", "Hoja_Origen"] + [col for col in original_columns if col not in ["Origen_Archivo", "Dato_Buscado", "Hoja_Origen"]]
                matching_rows = matching_rows.select(columns_order)
                return matching_rows
            return None
            
        except Exception as e:
            print(f"❌ Search error in {filepath} (Sheet: {sheet_name}): {e}")
            return search_in_dataframe_fallback(df_pl, value, filepath, sheet_name)

    def search_in_dataframe_fallback(df_pl, value, filepath, sheet_name="N/A"):
        """Alternative search method"""
        try:
            df_str = df_pl.cast(pl.Utf8, strict=False).fill_null("")
            
            # Use fold for better performance with contains
            mask = pl.fold(
                acc=pl.lit(False),
                function=lambda acc, col: acc | col.str.contains(value, literal=True),
                exprs=[pl.col(col) for col in df_str.columns]
            )
            
            matching_rows = df_str.filter(mask)
            
            if len(matching_rows) > 0:
                original_columns = df_str.columns
                matching_rows = matching_rows.with_columns([
                    pl.lit(filepath).cast(pl.Utf8).alias("Origen_Archivo"),
                    pl.lit(value).cast(pl.Utf8).alias("Dato_Buscado"),
                    pl.lit(sheet_name).cast(pl.Utf8).alias("Hoja_Origen")
                ])
                columns_order = ["Origen_Archivo", "Dato_Buscado", "Hoja_Origen"] + [col for col in original_columns if col not in ["Origen_Archivo", "Dato_Buscado", "Hoja_Origen"]]
                matching_rows = matching_rows.select(columns_order)
                return matching_rows
            return None
            
        except Exception as e:
            print(f"❌ Fallback search also failed for {filepath} (Sheet: {sheet_name}): {e}")
            return search_in_dataframe_simple(df_pl, value, filepath, sheet_name)

    def search_in_dataframe_simple(df_pl, value, filepath, sheet_name="N/A"):
        """Simple row-by-row search as last resort"""
        try:
            df_pd = df_pl.to_pandas()
            df_pd = df_pd.astype(str).fillna('')
            
            # Use contains for partial matching
            mask = df_pd.apply(lambda row: any(str(value) in str(cell) for cell in row), axis=1)
            matching_rows = df_pd[mask].copy()
            
            if len(matching_rows) > 0:
                matching_rows.insert(0, 'Origen_Archivo', filepath)
                matching_rows.insert(1, 'Dato_Buscado', value)
                matching_rows.insert(2, 'Hoja_Origen', sheet_name)
                return pl.from_pandas(matching_rows)
            return None
            
        except Exception as e:
            print(f"❌ Simple search failed for {filepath} (Sheet: {sheet_name}): {e}")
            return None

    # --- Function to create search summary ---
    def create_search_summary(results, search_stats):
        """Create a DataFrame with search summary"""
        summary_data = []
        
        for value in search_stats['search_values']:
            found = value in results and len(results[value]) > 0
            match_count = sum(len(df) for df in results[value]) if found else 0
            file_count = len(results[value]) if found else 0
            
            summary_data.append({
                'Dato_Buscado': value,
                'Encontrado': 'SI' if found else 'NO',
                'Total_Coincidencias': match_count,
                'Archivos_Con_Coincidencias': file_count,
                'Estado': 'ENCONTRADO' if found else 'NO_ENCONTRADO'
            })
        
        # General statistics
        summary_data.extend([
            {'Dato_Buscado': '=== RESUMEN ===', 'Encontrado': '', 'Total_Coincidencias': '', 'Archivos_Con_Coincidencias': '', 'Estado': ''},
            {'Dato_Buscado': 'Total valores buscados', 'Encontrado': len(search_stats['search_values']), 'Total_Coincidencias': '', 'Archivos_Con_Coincidencias': '', 'Estado': ''},
            {'Dato_Buscado': 'Valores encontrados', 'Encontrado': len([v for v in search_stats['search_values'] if v in results]), 'Total_Coincidencias': '', 'Archivos_Con_Coincidencias': '', 'Estado': ''},
            {'Dato_Buscado': 'Valores no encontrados', 'Encontrado': len([v for v in search_stats['search_values'] if v not in results]), 'Total_Coincidencias': '', 'Archivos_Con_Coincidencias': '', 'Estado': ''},
            {'Dato_Buscado': 'Total coincidencias', 'Encontrado': '', 'Total_Coincidencias': search_stats['total_matches'], 'Archivos_Con_Coincidencias': '', 'Estado': ''},
            {'Dato_Buscado': 'Archivos con coincidencias', 'Encontrado': '', 'Total_Coincidencias': '', 'Archivos_Con_Coincidencias': search_stats['files_with_matches'], 'Estado': ''},
            {'Dato_Buscado': 'Archivos con errores', 'Encontrado': '', 'Total_Coincidencias': '', 'Archivos_Con_Coincidencias': len(search_stats['files_with_errors']), 'Estado': ''}
        ])
        
        return pl.DataFrame(summary_data)

    # --- Function to align DataFrame schemas dynamically ---
    def align_dataframe_schemas(dataframes):
        """Align all DataFrames to have the same column structure - preserving original structure"""
        if not dataframes:
            return []
        
        # Get all unique columns from all DataFrames, but preserve order from first DataFrame
        all_columns = set()
        column_order = []
        
        # First, collect original column order from the first dataframe (excluding metadata columns)
        if dataframes and len(dataframes) > 0:
            first_df = dataframes[0]
            metadata_cols = {"Origen_Archivo", "Dato_Buscado", "Hoja_Origen"}
            original_cols = [col for col in first_df.columns if col not in metadata_cols]
            column_order = original_cols
        
        # Then collect all unique columns from all dataframes
        for df in dataframes:
            for col in df.columns:
                if col not in metadata_cols:  # Exclude metadata columns from dynamic collection
                    all_columns.add(col)
        
        # If we have column order from first DF, use it as base, then add any new columns
        if column_order:
            # Add any columns that weren't in the first DF but appear in others
            extra_cols = sorted([col for col in all_columns if col not in column_order])
            final_column_order = column_order + extra_cols
        else:
            final_column_order = sorted(list(all_columns))
        
        # Add metadata columns at the beginning
        final_column_order = ["Origen_Archivo", "Dato_Buscado", "Hoja_Origen"] + final_column_order
        
        aligned_dataframes = []
        for df in dataframes:
            try:
                # First, ensure all existing columns are string type
                df = df.cast({col: pl.Utf8 for col in df.columns})
                
                # Add missing columns with empty string values (not null)
                missing_columns = [col for col in final_column_order if col not in df.columns]
                if missing_columns:
                    for col in missing_columns:
                        df = df.with_columns(pl.lit("").alias(col))
                
                # Reorder columns to match the final schema
                df = df.select(final_column_order)
                aligned_dataframes.append(df)
                
            except Exception as e:
                print(f"⚠️ Error aligning DataFrame schema: {e}")
                # Fallback: convert to pandas and back
                try:
                    df_pd = df.to_pandas()
                    df_pd = df_pd.astype(str).fillna('')
                    # Add missing columns
                    for col in missing_columns:
                        df_pd[col] = ""
                    # Reorder columns
                    df_pd = df_pd[final_column_order]
                    aligned_dataframes.append(pl.from_pandas(df_pd))
                except Exception as pd_error:
                    print(f"❌ Could not align DataFrame: {pd_error}")
                    continue
        
        return aligned_dataframes

    # --- Robust function to combine DataFrames ---
    def combine_dataframes_safely(dataframes):
        """Safely combine DataFrames with robust error handling"""
        if not dataframes:
            return None
        
        if len(dataframes) == 1:
            return dataframes[0]
        
        # Try multiple combination strategies
        strategies = [
            # Strategy 1: Direct concatenation after alignment
            lambda: pl.concat(align_dataframe_schemas(dataframes)),
            
            # Strategy 2: Convert all to string and concatenate
            lambda: pl.concat([df.cast(pl.Utf8).fill_null("") for df in dataframes]),
            
            # Strategy 3: Use pandas as intermediate
            lambda: pl.from_pandas(pd.concat([df.to_pandas() for df in dataframes], ignore_index=True)),
            
            # Strategy 4: Manual combination
            lambda: manual_dataframe_combination(dataframes)
        ]
        
        for i, strategy in enumerate(strategies, 1):
            try:
                print(f"🔄 Trying combination strategy {i}...")
                result = strategy()
                if result is not None and len(result) > 0:
                    print(f"✅ Success with strategy {i}")
                    return result
            except Exception as e:
                print(f"❌ Strategy {i} failed: {e}")
                continue
        
        print("💥 All combination strategies failed")
        return None

    def manual_dataframe_combination(dataframes):
        """Manual combination as last resort"""
        try:
            # Convert all to pandas and combine
            pandas_dfs = []
            for df in dataframes:
                df_pd = df.to_pandas()
                df_pd = df_pd.astype(str).fillna('')
                pandas_dfs.append(df_pd)
            
            combined_pd = pd.concat(pandas_dfs, ignore_index=True)
            return pl.from_pandas(combined_pd)
        except Exception as e:
            print(f"❌ Manual combination failed: {e}")
            return None

    # --- Function to export results ---
    def export_results(results, search_stats, output_path, base_filename):
        """Export results trying Excel first, then CSV, then JSON"""
        
        if not results:
            print("❗ No results to export.")
            return None
        
        # Create consolidated DataFrame with all results
        all_results = []
        for value, df_list in results.items():
            if df_list:
                all_results.extend(df_list)
        
        if not all_results:
            print("❗ No data to export.")
            return None
        
        # Use safe combination instead of direct concatenation
        print("🔄 Combining DataFrames safely...")
        combined_pl = combine_dataframes_safely(all_results)
        
        if combined_pl is None:
            print("❌ Could not combine DataFrames")
            return None
        
        # Create summary
        summary_pl = create_search_summary(results, search_stats)
        
        print(f"💾 Exporting {len(combined_pl)} rows...")
        
        # Try 1: Export to Excel
        try:
            excel_path = os.path.join(output_path, f"{base_filename}.xlsx")
            print("📗 Attempting Excel export...")
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # DETAIL_SEARCH sheet with all results (ORGANIZED)
                combined_pl.to_pandas().to_excel(writer, sheet_name="DETALLE_BUSQUEDA", index=False)
                
                # SEARCH_SUMMARY sheet with summary
                summary_pl.to_pandas().to_excel(writer, sheet_name="RESUMEN_BUSQUEDA", index=False)
                
                # Separate sheets by search value
                for value, df_list in results.items():
                    if df_list:
                        # Combine value-specific dataframes safely
                        value_combined = combine_dataframes_safely(df_list)
                        if value_combined is not None and len(value_combined) > 0:
                            sheet_name = f"VALOR_{str(value)[:25]}"
                            sheet_name = ''.join(c for c in sheet_name if c.isalnum() or c in (' ', '_', '-')).strip()
                            if not sheet_name:
                                sheet_name = f"Valor_{hash(value) % 10000:04d}"
                            # Limit sheet name to 31 characters (Excel limit)
                            sheet_name = sheet_name[:31]
                            value_combined.to_pandas().to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f"✅ Successfully exported to Excel: {excel_path}")
            return excel_path
            
        except Exception as e:
            print(f"❌ Excel export failed: {e}")
            print("📘 Attempting CSV export...")
            
            # Try 2: Export to CSV (multiple files)
            try:
                # Main file with all results
                main_csv_path = os.path.join(output_path, f"{base_filename}_DETALLE_BUSQUEDA.csv")
                combined_pl.write_csv(main_csv_path)
                
                # Summary file
                summary_csv_path = os.path.join(output_path, f"{base_filename}_RESUMEN_BUSQUEDA.csv")
                summary_pl.write_csv(summary_csv_path)
                
                # Files by search value
                for value, df_list in results.items():
                    if df_list:
                        value_combined = combine_dataframes_safely(df_list)
                        if value_combined is not None and len(value_combined) > 0:
                            safe_value_name = "".join(c for c in str(value) if c.isalnum() or c in (' ', '_', '-')).strip()
                            if not safe_value_name:
                                safe_value_name = f"Valor_{hash(value) % 10000:04d}"
                            value_csv_path = os.path.join(output_path, f"{base_filename}_VALOR_{safe_value_name[:50]}.csv")
                            value_combined.write_csv(value_csv_path)
                
                print(f"✅ Successfully exported to CSV files in: {output_path}")
                return main_csv_path
                
            except Exception as e:
                print(f"❌ CSV export failed: {e}")
                print("📙 Attempting JSON export...")
                
                # Try 3: Export to JSON (structured)
                try:
                    json_path = os.path.join(output_path, f"{base_filename}.json")
                    
                    # Create organized JSON structure
                    json_data = {
                        "resumen_busqueda": summary_pl.to_dicts(),
                        "detalle_busqueda": combined_pl.to_dicts(),
                        "metadata_busqueda": {
                            "total_archivos_procesados": search_stats['total_files_processed'],
                            "archivos_con_coincidencias": search_stats['files_with_matches'],
                            "total_coincidencias": search_stats['total_matches'],
                            "fecha_hora_busqueda": datetime_now,
                            "archivos_con_errores": search_stats['files_with_errors']
                        }
                    }
                    
                    with open(json_path, 'w', encoding='utf-8') as f:
                        json.dump(json_data, f, indent=2, ensure_ascii=False)
                    
                    print(f"✅ Successfully exported to JSON: {json_path}")
                    return json_path
                    
                except Exception as e:
                    print(f"❌ JSON export failed: {e}")
                    
                    # Last attempt: simple JSON
                    try:
                        print("🔄 Attempting simple JSON export...")
                        simple_json_path = os.path.join(output_path, f"{base_filename}_simple.json")
                        combined_pl.write_ndjson(simple_json_path)
                        print(f"✅ Successfully exported to simple JSON: {simple_json_path}")
                        return simple_json_path
                        
                    except Exception as final_error:
                        print(f"💥 All export methods failed: {final_error}")
                        return None
    
    # --- FILE SEARCH ---
    
    # Find files recursively - supporting both .parquet and .parquets extensions
    print("🔍 Searching for files in directory and subdirectories...")
    
    csv_files = find_files_recursive(directory, ['.csv'])
    parquet_files = find_files_recursive(directory, ['.parquet', '.parquets'])  # Support both extensions
    excel_files = find_files_recursive(directory, ['.xlsx', '.xls'])
    
    print(f"📊 Found {len(csv_files)} CSV files")
    print(f"⚡ Found {len(parquet_files)} Parquet files")
    print(f"📑 Found {len(excel_files)} Excel files")
    
    total_files = len(csv_files) + len(parquet_files) + len(excel_files)
    search_stats['total_files_processed'] = total_files
    print(f"📁 Total files to process: {total_files}")

    # --- Search CSV Files ---
    print(f"📊 Processing {len(csv_files)} CSV files...")
    
    for file_path in csv_files:
        filename = os.path.basename(file_path)
        relative_path = os.path.relpath(file_path, directory)
        file_size = os.path.getsize(file_path) / (1024 * 1024)
        
        search_stats['files_by_type']['csv'] += 1
        
        if file_size > 50:
            print(f"🚀 Using Polars for large CSV: {relative_path} ({file_size:.1f} MB)")
            df = read_csv_polars_optimized(file_path)
        else:
            df_pd = read_csv_dynamic(file_path)
            df = pl.from_pandas(df_pd) if df_pd is not None else None
        
        if df is None:
            print(f"⛔ Skipping unreadable CSV: {relative_path}")
            continue
        
        file_has_matches = False
        for value in search_values:
            matching_rows = search_in_dataframe_polars(df, value, relative_path, "CSV")
            
            if matching_rows is not None and len(matching_rows) > 0:
                print(f"✅ Valor '{value}' encontrado en CSV: {relative_path} ({len(matching_rows)} coincidencias)")
                
                if value not in results:
                    results[value] = []
                results[value].append(matching_rows)
                search_stats['total_matches'] += len(matching_rows)
                file_has_matches = True
                
                if value not in search_stats['matches_by_value']:
                    search_stats['matches_by_value'][value] = 0
                search_stats['matches_by_value'][value] += len(matching_rows)
        
        if file_has_matches:
            search_stats['files_with_matches'] += 1

    # --- Search Parquet Files ---
    print(f"⚡ Processing {len(parquet_files)} Parquet files...")
    
    for file_path in parquet_files:
        filename = os.path.basename(file_path)
        relative_path = os.path.relpath(file_path, directory)
        
        search_stats['files_by_type']['parquet'] += 1
        
        df = read_parquet_optimized(file_path)
        
        if df is None:
            print(f"⛔ Skipping unreadable Parquet: {relative_path}")
            continue
        
        file_has_matches = False
        for value in search_values:
            matching_rows = search_in_dataframe_polars(df, value, relative_path, "Parquet")
            if matching_rows is not None and len(matching_rows) > 0:
                print(f"✅ Valor '{value}' encontrado en Parquet: {relative_path} ({len(matching_rows)} coincidencias)")
                
                if value not in results:
                    results[value] = []
                results[value].append(matching_rows)
                search_stats['total_matches'] += len(matching_rows)
                file_has_matches = True
                
                if value not in search_stats['matches_by_value']:
                    search_stats['matches_by_value'][value] = 0
                search_stats['matches_by_value'][value] += len(matching_rows)
        
        if file_has_matches:
            search_stats['files_with_matches'] += 1

    # --- Search Excel Files ---
    print(f"📊 Processing {len(excel_files)} Excel files...")
    
    for file_path in excel_files:
        filename = os.path.basename(file_path)
        relative_path = os.path.relpath(file_path, directory)
        
        search_stats['files_by_type']['excel'] += 1
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            print(f"📑 Archivo Excel '{relative_path}' tiene {len(sheet_names)} hojas")
            
        except Exception as e:
            print(f"❌ No se pudo leer el archivo Excel {relative_path}: {e}")
            search_stats['files_with_errors'].append(f"{relative_path} - Excel error: {e}")
            continue

        file_has_matches = False
        for sheet_name in sheet_names:
            try:
                df_pd = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, engine='openpyxl')
                df_pd = df_pd.fillna('')
                
                if df_pd.empty:
                    continue
                    
                df = pl.from_pandas(df_pd)

                for value in search_values:
                    matching_rows = search_in_dataframe_polars(df, value, relative_path, sheet_name)
                    if matching_rows is not None and len(matching_rows) > 0:
                        print(f"✅ Valor '{value}' encontrado en Excel: {relative_path} (Hoja: {sheet_name}) - {len(matching_rows)} coincidencias")
                        
                        if value not in results:
                            results[value] = []
                        results[value].append(matching_rows)
                        search_stats['total_matches'] += len(matching_rows)
                        file_has_matches = True
                        
                        if value not in search_stats['matches_by_value']:
                            search_stats['matches_by_value'][value] = 0
                        search_stats['matches_by_value'][value] += len(matching_rows)
                        
            except Exception as e:
                print(f"❌ No se pudo leer la hoja '{sheet_name}' de {relative_path}: {e}")
                continue
        
        if file_has_matches:
            search_stats['files_with_matches'] += 1

    # --- FINAL RESULTS ---
    print("\n📈 RESUMEN DE BÚSQUEDA:")
    print("=" * 60)
    for value in search_values:
        if value in results and len(results[value]) > 0:
            total_matches = sum(len(df) for df in results[value])
            print(f"✅ '{value}': {total_matches} coincidencias en {len(results[value])} archivos")
        else:
            print(f"❌ '{value}': No se encontraron coincidencias")
    
    print("=" * 60)
    print(f"📊 TOTAL: {search_stats['total_matches']} coincidencias en {search_stats['files_with_matches']} archivos")
    print(f"📁 Procesados: {search_stats['total_files_processed']} archivos")
    print(f"⚠️ Errores: {len(search_stats['files_with_errors'])} archivos")

    if not results:
        print("❗ No se encontraron coincidencias en ningún archivo. Nada que exportar.")
        return None, search_stats

    # Export results
    print(f"\n💾 Exportando resultados a: {base_output_path}")
    exported_file = export_results(results, search_stats, base_output_path, base_filename)
    
    if exported_file:
        print(f"📁 Resultados exportados exitosamente a: {exported_file}")
        print(f"🎯 ¡Búsqueda completada! Total de coincidencias: {search_stats['total_matches']}")
    else:
        print("❌ Falló la exportación de resultados usando todos los métodos.")
    
    return results, search_stats