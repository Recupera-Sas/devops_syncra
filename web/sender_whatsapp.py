import random
from openpyxl import Workbook, load_workbook
from urllib.parse import quote_plus
import os
from selenium.common.exceptions import TimeoutException
from PyQt6.QtWidgets import QMessageBox # This import is not used in the provided code, but kept as is.
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import urllib.parse # Already imported as urllib.parse, but kept as is.
import time
from selenium.webdriver.chrome.webdriver import WebDriver # This import is not used in the provided code, but kept as is.
from selenium.common.exceptions import WebDriverException
from selenium.common import exceptions as selexceptions # This import is not used in the provided code, but kept as is.
from selenium.webdriver.common.keys import Keys
import re
import csv # This import is not used in the provided code, but kept as is.
from datetime import datetime
import pandas as pd # Replaced PySpark with Pandas
import sys # This import is not used in the provided code, but kept as is.

# Removed Spark session initialization as we are using Pandas now.
# spark = get_spark_session() # Removed

def generate_message_column(df, template):
    """
    Generates a 'MESSAGE' column in the DataFrame based on a template string.
    The template can contain placeholders like (COLUMN_NAME) which will be replaced
    with the corresponding column values from the DataFrame.

    Args:
        df (pd.DataFrame): The input DataFrame.
        template (str): The template string for generating messages.

    Returns:
        pd.DataFrame: The DataFrame with the new 'MESSAGE' column, or None if a column is missing.
    """
    # Get column names and convert to uppercase for case-insensitive matching
    columns = [c.upper() for c in df.columns]
    # Split the template into text parts and variable placeholders (e.g., (NOMBRE))
    parts = re.split(r"(\([^)]+\))", template)
    
    # Initialize a list to hold the components for concatenation
    message_components = []

    for part in parts:
        if part.startswith("(") and part.endswith(")"):
            # Extract column name from placeholder (e.g., "NOMBRE" from "(NOMBRE)")
            col_name = part[1:-1].upper()
            if col_name not in columns:
                print(f"⚠️ Column '{col_name}' does not exist in the DataFrame.")
                return None
            # Append the DataFrame series for the column
            message_components.append(df[col_name])
        else:
            # Append the literal text part
            if part.strip() != "":
                message_components.append(pd.Series([part] * len(df))) # Create a Series of the literal part

    # Concatenate all parts (Series and literal Series) into the 'MESSAGE' column
    # This uses a loop to concatenate elements, ensuring correct string joining.
    df["MESSAGE"] = ""
    for component in message_components:
        # Ensure component is a Series for concatenation
        if isinstance(component, pd.Series):
            df["MESSAGE"] = df["MESSAGE"].astype(str) + component.astype(str)
        else:
            df["MESSAGE"] = df["MESSAGE"].astype(str) + str(component)

    return df

def read_csv_dynamic_encoding(selected_file, sep=';'):
    encodings = ["utf-8", "latin1", "cp1252"]
    for enc in encodings:
        try:
            df = pd.read_csv(selected_file, sep=sep, encoding=enc)
            print(f"Archivo leído correctamente con encoding: {enc}")
            return df
        except UnicodeDecodeError:
            print(f"Fallo con encoding: {enc}, probando otro...")
    raise Exception("No se pudo leer el archivo con los encodings probados.")

def read_file(selected_file, template):
    """
    Reads a CSV file into a Pandas DataFrame, converts column names to uppercase,
    and generates the 'MESSAGE' column based on the provided template.
    It also prints the first message and returns the DataFrame and a response.

    Args:
        selected_file (str): The path to the CSV file.
        template (str): The template string for generating messages.

    Returns:
        tuple: A tuple containing the DataFrame and a boolean response.
    """
    try:
        # Convert column names to uppercase
        df = read_csv_dynamic_encoding(selected_file, sep=';')
        df.columns = [col_name.upper() for col_name in df.columns]
        
        # Generate the message column
        df = generate_message_column(df, template)

        if df is None:
            # If generate_message_column returns None, it means a column was missing
            print("Failed to generate messages due to missing columns.")
            return None, False

        # Get the first message to be sent
        first_message = df["MESSAGE"].iloc[0]
        
        print("First message to be sent: ", first_message)

        response = True # This was a hardcoded True in the original, so keeping it.
        
        print("Response from user: ", response)
        
        return df, response
    except FileNotFoundError:
        print(f"Error: The file '{selected_file}' was not found.")
        return None, False
    except Exception as e:
        print(f"Error reading file or processing data: {e}")
        return None, False

def send_messages(selected_file, output_file, template, process_data):
    
    df, response = read_file(selected_file, template)
    
    if df is None or response == False:
        Message = "El archivo NO ha sido enviado debido a errores en la lectura o procesamiento de datos."
        return Message

    driver = None # Initialize driver to None for proper cleanup in case of early errors
    try:
        # Configure the browser once
        chrome_options = Options()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        # Suppress the "DevTools listening on ws://" message
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

        driver.get("https://web.whatsapp.com")
        print("🔒 Waiting for you to log in to WhatsApp Web...")
        # Wait until the side pane (indicating successful login) is present
        WebDriverWait(driver, 100).until(
            EC.presence_of_element_located((By.ID, "pane-side"))
        )
        print("✅ Logged in successfully.")

        # Extract numbers and messages from the Pandas DataFrame
        if "DATO_CONTACTO" in df.columns:
            numbers = df["DATO_CONTACTO"].tolist()
        else:
            numbers = df["CELULAR"].tolist()
        
        messages = df["MESSAGE"].tolist()
        
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S") # Kept as is, though not used directly in this loop
        date_for_file = datetime.now().strftime("%Y-%m-%d") # Kept as is, though not used directly in this loop
        status = None

    except Exception as e:
        Message = f"❌ Error durante WhatsApp Web setup o preparacion de data: {e}"
        print(Message)
        if driver:
            driver.quit()
        return Message
    
    counter_register = 0 # Counter for processed records
    for number, message in zip(numbers, messages):
        counter_register += 1
        left_messages = len(messages) - counter_register
        left_messages_to_send = f"😁 Procesando registro {counter_register} de {left_messages} restantes."
        print(left_messages_to_send)
        number = int(f"57{number}")
        encoded_message = quote_plus(message)
        url = f"https://web.whatsapp.com/send?phone={number}&text={encoded_message}"
        print(f"📨 Sending to {number} -> {message}")
        driver.get(url)
        
        status = None
        status = dynamic_send_messages(driver, number)

        # Save the status of each message to an Excel report
        save_to_excel(output_file, number, message, status)
        
        number = int(f"573118025363")
        url = f"https://web.whatsapp.com/send?phone={number}&text={left_messages_to_send}"
        driver.get(url)
        dynamic_send_messages(driver, number)

    # Close the browser after all messages are processed
    if driver:
        driver.quit()
    Message = "El archivo ha sido tratado exitosamente y su reporte esta en las Descargas."
    
    return Message

def dynamic_send_messages(driver, number):
    
    try:
        # This block attempts to click a button that might appear before logging in.
        # It's based on the original code's logic.
        try:
            random_wait = random.uniform(2, 4)
            # XPath for a common "Continue" or "OK" button that might pop up
            button = WebDriverWait(driver, random_wait).until(EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/span[2]/div/div/div/div/div/div/div[2]/div/button')))
            button.click()
            print("⚠️ Clicked a preliminary button (if present).")
        except TimeoutException:
            pass # If the button isn't found, we just continue
        
        print("Entering message.")
        try:
            time.sleep(1)
            print("In message box.")
            
            random_wait = random.uniform(3, 15)
            if number == 573180945484:
                random_wait = 1
                
            try:
                send_button = WebDriverWait(driver, random_wait).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div/div[4]/div/span/div/div/div[1]/div[1]/span'))
                )
                
                # send_image(driver, number)
                
                send_button.click()
                print(f"✅ Enviar {random_wait} segundos")
                status = "Enviado"
                
            except WebDriverException as e:
                try:
                    
                    send_button = WebDriverWait(driver, random_wait).until(
                        EC.element_to_be_clickable((By.XPATH, '//button[@aria-label="Enviar"]'))
                    )
                    
                    # send_image(driver, number)
                    
                    send_button.click()
                    print(f"✅ Enviar {random_wait} segundos")
                    status = "Enviado"
                except WebDriverException as e:
                    send_button = WebDriverWait(driver, random_wait).until(
                        EC.element_to_be_clickable((By.XPATH, '//button[@aria-label="Send"]'))
                    )
                    
                    # send_image(driver, number)
                    
                    send_button.click()
                    print(f"✅ Send {random_wait} seconds")
                    status = "Enviado"
            
            time.sleep(random_wait) # Short delay after sending
            
        except WebDriverException as e:
            print(f"❌ Error al intentar hacer clic en el botón")
            status = "Error"

    except Exception as e:
        status = "No enviado"
        print(f"❌ Error with number {number} error: {e}")
        
    return status

def send_image(driver, number):
    
    if number != 573180945484:
        
        status = "No enviado"

        try:
            print("Iniciando el proceso para pegar y enviar la imagen desde el portapapeles...")

            # Esperar a que el campo de texto del chat esté presente y hacer clic para enfocarlo.
            # Esto es crucial para que la acción de pegar funcione.
            chat_box = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]'))
            )
            chat_box.click()
            print("✅ Se hizo clic en el cuadro de texto del chat para enfocarlo.")

            # Usar ActionChains para simular la combinación de teclas Ctrl+V (pegar).
            actions = ActionChains(driver)
            # La combinación de teclas es diferente para cada sistema operativo.
            # Keys.CONTROL funciona en la mayoría de los casos para Windows y Linux.
            # Para macOS, se usaría Keys.COMMAND.
            actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            print("✅ Se simuló la acción de pegar (Ctrl+V).")
            
            # Esperar a que aparezca el botón de enviar después de pegar la imagen.
            send_media_button = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[1]/div[3]/div/div[2]/div[2]/span/div/div/div/div[2]/div/div[2]/div[2]/div/div/span'))
            )
            send_media_button.click()
            print("✅ Imagen pegada y enviada correctamente.")
            status = "Enviado"

        except TimeoutException:
            print("❌ Error de tiempo de espera: No se pudo encontrar un elemento.")
            status = "Error de tiempo de espera"
        except WebDriverException as e:
            print(f"❌ Error de WebDriver: {e}")
            status = "Error de WebDriver"
        except Exception as e:
            print(f"❌ Ocurrió un error inesperado: {e}")
            status = "Error inesperado"
    
    else:
        status = None
        pass
    
    return status

def save_to_excel(output_folder, number, message, status):
    """
    Saves the message sending status to an Excel file.
    It creates a new file daily or appends to an existing one.

    Args:
        output_folder (str): The folder where the Excel file will be saved.
        number (str): The phone number to which the message was sent.
        message (str): The content of the message sent.
        status (str): The sending status ("Enviado" or "No enviado").
    """
    # Create the file name with the current date
    date_for_file = datetime.now().strftime("%Y-%m-%d")
    file_path = os.path.join(output_folder, f"Reporte RPA WhatsApp {date_for_file}.xlsx")

    # Try to load the workbook if it already exists, or create a new one
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Write headers only if the file is new
        sheet.append(["Número", "Mensaje", "Fecha Hora", "Estado"])

    # Add row with data
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([number, message, current_datetime, status])

    # Save the file
    workbook.save(file_path)